"""
Full EDA Script – Customer Data
Generates: charts (PNG), Excel report, PDF report
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
import warnings, os
warnings.filterwarnings('ignore')

# ── paths ────────────────────────────────────────────────────────────────────
DATA  = '/mnt/project/Customer_Data.csv'
OUT   = '/mnt/user-data/outputs'
IMG   = '/home/claude/charts'
os.makedirs(OUT, exist_ok=True)
os.makedirs(IMG, exist_ok=True)

# ── palette ───────────────────────────────────────────────────────────────────
COLORS = ['#2563EB','#10B981','#F59E0B','#EF4444','#8B5CF6','#EC4899','#06B6D4']
sns.set_theme(style='whitegrid', palette=COLORS)

df = pd.read_csv(DATA)
df_churned = df[df['Customer_Status'] == 'Churned']
churn_rate = round(len(df_churned)/len(df)*100, 1)

# ════════════════════════════════════════════════════════════════════════════
# 1. CUSTOMER STATUS DONUT
# ════════════════════════════════════════════════════════════════════════════
fig, ax = plt.subplots(figsize=(7, 5))
status_counts = df['Customer_Status'].value_counts()
wedges, texts, autotexts = ax.pie(
    status_counts, labels=status_counts.index, autopct='%1.1f%%',
    colors=COLORS[:3], startangle=140, pctdistance=0.78,
    wedgeprops=dict(width=0.55, edgecolor='white', linewidth=2))
for t in autotexts: t.set_fontsize(11); t.set_fontweight('bold')
ax.set_title('Customer Status Distribution', fontsize=14, fontweight='bold', pad=20)
plt.tight_layout()
plt.savefig(f'{IMG}/01_status_donut.png', dpi=150, bbox_inches='tight')
plt.close()

# ════════════════════════════════════════════════════════════════════════════
# 2. CHURN CATEGORY BAR
# ════════════════════════════════════════════════════════════════════════════
fig, ax = plt.subplots(figsize=(8, 5))
cc = df_churned['Churn_Category'].value_counts()
bars = ax.barh(cc.index, cc.values, color=COLORS[:len(cc)], edgecolor='white', height=0.6)
for bar in bars:
    ax.text(bar.get_width()+8, bar.get_y()+bar.get_height()/2,
            f'{bar.get_width():,}', va='center', fontsize=10, fontweight='bold')
ax.set_xlabel('Number of Customers', fontsize=11)
ax.set_title('Churn by Category', fontsize=14, fontweight='bold')
ax.set_xlim(0, cc.max()*1.15)
plt.tight_layout()
plt.savefig(f'{IMG}/02_churn_category.png', dpi=150, bbox_inches='tight')
plt.close()

# ════════════════════════════════════════════════════════════════════════════
# 3. AGE DISTRIBUTION BY STATUS
# ════════════════════════════════════════════════════════════════════════════
fig, ax = plt.subplots(figsize=(9, 5))
for i, (status, color) in enumerate(zip(['Stayed','Churned','Joined'], COLORS)):
    subset = df[df['Customer_Status']==status]['Age']
    ax.hist(subset, bins=25, alpha=0.65, color=color, label=status, edgecolor='white')
ax.set_xlabel('Age', fontsize=11); ax.set_ylabel('Count', fontsize=11)
ax.set_title('Age Distribution by Customer Status', fontsize=14, fontweight='bold')
ax.legend(fontsize=10)
plt.tight_layout()
plt.savefig(f'{IMG}/03_age_distribution.png', dpi=150, bbox_inches='tight')
plt.close()

# ════════════════════════════════════════════════════════════════════════════
# 4. MONTHLY CHARGE vs TENURE (scatter)
# ════════════════════════════════════════════════════════════════════════════
fig, ax = plt.subplots(figsize=(9, 6))
status_map = {'Stayed': COLORS[1], 'Churned': COLORS[3], 'Joined': COLORS[0]}
for status, color in status_map.items():
    sub = df[df['Customer_Status']==status]
    ax.scatter(sub['Tenure_in_Months'], sub['Monthly_Charge'],
               c=color, alpha=0.35, s=18, label=status)
ax.set_xlabel('Tenure (Months)', fontsize=11)
ax.set_ylabel('Monthly Charge ($)', fontsize=11)
ax.set_title('Monthly Charge vs Tenure by Status', fontsize=14, fontweight='bold')
ax.legend(fontsize=10)
plt.tight_layout()
plt.savefig(f'{IMG}/04_charge_vs_tenure.png', dpi=150, bbox_inches='tight')
plt.close()

# ════════════════════════════════════════════════════════════════════════════
# 5. CONTRACT TYPE CHURN RATE
# ════════════════════════════════════════════════════════════════════════════
fig, ax = plt.subplots(figsize=(8, 5))
contract_churn = df.groupby('Contract').apply(
    lambda x: (x['Customer_Status']=='Churned').sum()/len(x)*100).reset_index()
contract_churn.columns = ['Contract', 'Churn_Rate']
contract_churn = contract_churn.sort_values('Churn_Rate', ascending=False)
bars = ax.bar(contract_churn['Contract'], contract_churn['Churn_Rate'],
              color=COLORS[:3], edgecolor='white', width=0.5)
for bar in bars:
    ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.5,
            f'{bar.get_height():.1f}%', ha='center', fontsize=11, fontweight='bold')
ax.set_ylabel('Churn Rate (%)', fontsize=11)
ax.set_title('Churn Rate by Contract Type', fontsize=14, fontweight='bold')
ax.set_ylim(0, contract_churn['Churn_Rate'].max()*1.15)
plt.tight_layout()
plt.savefig(f'{IMG}/05_contract_churn.png', dpi=150, bbox_inches='tight')
plt.close()

# ════════════════════════════════════════════════════════════════════════════
# 6. TOP CHURN REASONS
# ════════════════════════════════════════════════════════════════════════════
fig, ax = plt.subplots(figsize=(9, 6))
top_reasons = df_churned['Churn_Reason'].value_counts().head(10)
colors_grad = sns.color_palette('Blues_r', len(top_reasons))
bars = ax.barh(top_reasons.index[::-1], top_reasons.values[::-1],
               color=colors_grad[::-1], edgecolor='white', height=0.6)
for bar in bars:
    ax.text(bar.get_width()+3, bar.get_y()+bar.get_height()/2,
            f'{bar.get_width():,}', va='center', fontsize=10)
ax.set_xlabel('Number of Customers', fontsize=11)
ax.set_title('Top 10 Churn Reasons', fontsize=14, fontweight='bold')
ax.set_xlim(0, top_reasons.max()*1.15)
plt.tight_layout()
plt.savefig(f'{IMG}/06_churn_reasons.png', dpi=150, bbox_inches='tight')
plt.close()

# ════════════════════════════════════════════════════════════════════════════
# 7. INTERNET TYPE vs CHURN
# ════════════════════════════════════════════════════════════════════════════
fig, ax = plt.subplots(figsize=(8, 5))
internet_churn = df[df['Internet_Service']=='Yes'].groupby('Internet_Type').apply(
    lambda x: (x['Customer_Status']=='Churned').sum()/len(x)*100).reset_index()
internet_churn.columns = ['Internet_Type','Churn_Rate']
bars = ax.bar(internet_churn['Internet_Type'], internet_churn['Churn_Rate'],
              color=COLORS[:3], edgecolor='white', width=0.45)
for bar in bars:
    ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.5,
            f'{bar.get_height():.1f}%', ha='center', fontsize=11, fontweight='bold')
ax.set_ylabel('Churn Rate (%)', fontsize=11)
ax.set_title('Churn Rate by Internet Type', fontsize=14, fontweight='bold')
ax.set_ylim(0, internet_churn['Churn_Rate'].max()*1.15)
plt.tight_layout()
plt.savefig(f'{IMG}/07_internet_churn.png', dpi=150, bbox_inches='tight')
plt.close()

# ════════════════════════════════════════════════════════════════════════════
# 8. REVENUE COMPARISON (Box plots)
# ════════════════════════════════════════════════════════════════════════════
fig, axes = plt.subplots(1, 2, figsize=(11, 5))
for ax, col, title in zip(axes, ['Monthly_Charge','Total_Revenue'],
                          ['Monthly Charge ($)','Total Revenue ($)']):
    data_list = [df[df['Customer_Status']==s][col].dropna() for s in ['Stayed','Churned','Joined']]
    bp = ax.boxplot(data_list, patch_artist=True, notch=False,
                    medianprops=dict(color='black', linewidth=2))
    for patch, color in zip(bp['boxes'], [COLORS[1], COLORS[3], COLORS[0]]):
        patch.set_facecolor(color); patch.set_alpha(0.7)
    ax.set_xticklabels(['Stayed','Churned','Joined'], fontsize=10)
    ax.set_ylabel(title, fontsize=10)
    ax.set_title(f'{title} by Status', fontsize=12, fontweight='bold')
plt.tight_layout()
plt.savefig(f'{IMG}/08_revenue_boxplot.png', dpi=150, bbox_inches='tight')
plt.close()

# ════════════════════════════════════════════════════════════════════════════
# 9. CORRELATION HEATMAP
# ════════════════════════════════════════════════════════════════════════════
fig, ax = plt.subplots(figsize=(9, 7))
num_cols = ['Age','Tenure_in_Months','Monthly_Charge','Total_Charges',
            'Total_Revenue','Number_of_Referrals','Total_Long_Distance_Charges']
corr = df[num_cols].corr()
mask = np.triu(np.ones_like(corr, dtype=bool))
sns.heatmap(corr, mask=mask, annot=True, fmt='.2f', cmap='RdYlGn',
            center=0, ax=ax, linewidths=0.5, annot_kws={'size':9})
ax.set_title('Correlation Heatmap – Numeric Features', fontsize=14, fontweight='bold')
plt.tight_layout()
plt.savefig(f'{IMG}/09_correlation.png', dpi=150, bbox_inches='tight')
plt.close()

# ════════════════════════════════════════════════════════════════════════════
# 10. TOP 10 STATES BY CHURN
# ════════════════════════════════════════════════════════════════════════════
fig, ax = plt.subplots(figsize=(9, 6))
state_churn = df.groupby('State').apply(
    lambda x: (x['Customer_Status']=='Churned').sum()/len(x)*100
).sort_values(ascending=False).head(10).reset_index()
state_churn.columns = ['State','Churn_Rate']
bars = ax.bar(state_churn['State'], state_churn['Churn_Rate'],
              color=sns.color_palette('Reds_r', 10), edgecolor='white')
for bar in bars:
    ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.3,
            f'{bar.get_height():.1f}%', ha='center', fontsize=9, fontweight='bold')
ax.set_ylabel('Churn Rate (%)', fontsize=11)
ax.set_title('Top 10 States by Churn Rate', fontsize=14, fontweight='bold')
plt.xticks(rotation=30, ha='right')
plt.tight_layout()
plt.savefig(f'{IMG}/10_state_churn.png', dpi=150, bbox_inches='tight')
plt.close()

print("All charts saved ✓")


# ════════════════════════════════════════════════════════════════════════════
# EXCEL REPORT
# ════════════════════════════════════════════════════════════════════════════
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── helpers ──────────────────────────────────────────────────────────────────
BLUE  = PatternFill('solid', fgColor='2563EB')
LBLUE = PatternFill('solid', fgColor='DBEAFE')
GRAY  = PatternFill('solid', fgColor='F1F5F9')
WHITE = PatternFill('solid', fgColor='FFFFFF')
HDR   = Font(name='Arial', bold=True, color='FFFFFF', size=11)
SUBHDR= Font(name='Arial', bold=True, color='2563EB', size=10)
CELL  = Font(name='Arial', size=10)
BOLD  = Font(name='Arial', bold=True, size=10)
THIN  = Side(style='thin', color='CBD5E1')
def border(): return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
def hdr_row(ws, row, col, text, span=1, fill=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = HDR; c.fill = fill or BLUE
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border()
    if span > 1: ws.merge_cells(start_row=row, start_column=col,
                                end_row=row, end_column=col+span-1)
def write(ws, row, col, val, fmt=None, bold=False, fill=None, align='left'):
    c = ws.cell(row=row, column=col, value=val)
    c.font = BOLD if bold else CELL; c.border = border()
    c.fill = fill or WHITE
    c.alignment = Alignment(horizontal=align, vertical='center')
    if fmt: c.number_format = fmt

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: SUMMARY DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
ws1 = wb.active; ws1.title = 'Summary Dashboard'
ws1.column_dimensions['A'].width = 28
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 18
ws1.column_dimensions['E'].width = 18
ws1.column_dimensions['F'].width = 22
ws1.row_dimensions[1].height = 40

# Title
ws1.merge_cells('A1:F1')
t = ws1['A1']; t.value = 'Customer EDA Report – Summary Dashboard'
t.font = Font(name='Arial', bold=True, size=16, color='FFFFFF')
t.fill = PatternFill('solid', fgColor='1E3A5F')
t.alignment = Alignment(horizontal='center', vertical='center')

# KPI section
ws1.merge_cells('A3:F3')
k = ws1['A3']; k.value = 'KEY PERFORMANCE INDICATORS'
k.font = Font(name='Arial', bold=True, size=12, color='1E3A5F')
k.fill = LBLUE; k.alignment = Alignment(horizontal='center', vertical='center')
k.border = border(); ws1.row_dimensions[3].height = 24

kpis = [
    ('Total Customers', len(df), '#,##0'),
    ('Churned Customers', len(df_churned), '#,##0'),
    ('Overall Churn Rate', churn_rate/100, '0.0%'),
    ('Avg Monthly Charge', df['Monthly_Charge'].mean(), '$#,##0.00'),
    ('Avg Tenure (Months)', df['Tenure_in_Months'].mean(), '0.0'),
    ('Total Revenue', df['Total_Revenue'].sum(), '$#,##0'),
]
fills = [PatternFill('solid', fgColor=c) for c in
         ['EFF6FF','FEF3C7','FEE2E2','ECFDF5','F3E8FF','E0F2FE']]
for i, (label, val, fmt) in enumerate(kpis):
    col = i+1
    ws1.row_dimensions[4].height = 20; ws1.row_dimensions[5].height = 28
    lc = ws1.cell(row=4, column=col, value=label)
    lc.font = Font(name='Arial', bold=True, size=9, color='475569')
    lc.fill = fills[i]; lc.alignment = Alignment(horizontal='center')
    lc.border = border()
    vc = ws1.cell(row=5, column=col, value=val)
    vc.font = Font(name='Arial', bold=True, size=13, color='1E3A5F')
    vc.number_format = fmt; vc.fill = fills[i]
    vc.alignment = Alignment(horizontal='center', vertical='center')
    vc.border = border()

# Status breakdown table
row = 7
hdr_row(ws1, row, 1, 'Customer Status Breakdown', span=3)
row += 1
for h, col in zip(['Status','Count','Percentage'], [1,2,3]):
    hdr_row(ws1, row, col, h, fill=PatternFill('solid', fgColor='475569'))
row += 1
for i, (status, cnt) in enumerate(df['Customer_Status'].value_counts().items()):
    fill = GRAY if i%2 else WHITE
    write(ws1, row, 1, status, bold=True, fill=fill)
    write(ws1, row, 2, cnt, '#,##0', fill=fill, align='right')
    write(ws1, row, 3, cnt/len(df), '0.0%', fill=fill, align='right')
    row += 1

# Churn by category
row += 1
hdr_row(ws1, row, 1, 'Churn Category Breakdown', span=3)
row += 1
for h, col in zip(['Category','Count','% of Churned'], [1,2,3]):
    hdr_row(ws1, row, col, h, fill=PatternFill('solid', fgColor='475569'))
row += 1
for i, (cat, cnt) in enumerate(df_churned['Churn_Category'].value_counts().items()):
    fill = GRAY if i%2 else WHITE
    write(ws1, row, 1, cat, bold=True, fill=fill)
    write(ws1, row, 2, cnt, '#,##0', fill=fill, align='right')
    write(ws1, row, 3, cnt/len(df_churned), '0.0%', fill=fill, align='right')
    row += 1

# Contract churn rates
row += 1
hdr_row(ws1, row, 1, 'Contract Type Analysis', span=4)
row += 1
for h, col in zip(['Contract','Total','Churned','Churn Rate'], [1,2,3,4]):
    hdr_row(ws1, row, col, h, fill=PatternFill('solid', fgColor='475569'))
row += 1
for i, contract in enumerate(df['Contract'].unique()):
    sub = df[df['Contract']==contract]
    ch  = (sub['Customer_Status']=='Churned').sum()
    fill = GRAY if i%2 else WHITE
    write(ws1, row, 1, contract, bold=True, fill=fill)
    write(ws1, row, 2, len(sub), '#,##0', fill=fill, align='right')
    write(ws1, row, 3, ch, '#,##0', fill=fill, align='right')
    write(ws1, row, 4, ch/len(sub), '0.0%', fill=fill, align='right')
    row += 1

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: RAW DATA SUMMARY
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet('Data Overview')
ws2.column_dimensions['A'].width = 30
for c in 'BCDEFG': ws2.column_dimensions[c].width = 16
ws2.row_dimensions[1].height = 40

ws2.merge_cells('A1:G1')
t = ws2['A1']; t.value = 'Data Overview & Descriptive Statistics'
t.font = Font(name='Arial', bold=True, size=15, color='FFFFFF')
t.fill = PatternFill('solid', fgColor='1E3A5F')
t.alignment = Alignment(horizontal='center', vertical='center')

row = 3
hdr_row(ws2, row, 1, 'Dataset Info', span=2)
row += 1
for label, val in [('Total Records', len(df)), ('Total Columns', len(df.columns)),
                   ('Churn Rate', f"{churn_rate}%"),
                   ('Date Range (Tenure)', f"1 – {df['Tenure_in_Months'].max()} months")]:
    write(ws2, row, 1, label, bold=True, fill=GRAY)
    write(ws2, row, 2, val, fill=WHITE)
    row += 1

row += 1
hdr_row(ws2, row, 1, 'Descriptive Statistics – Numeric Columns', span=7)
row += 1
stats_hdrs = ['Column','Count','Mean','Std Dev','Min','Median','Max']
for col, h in enumerate(stats_hdrs, 1):
    hdr_row(ws2, row, col, h, fill=PatternFill('solid', fgColor='475569'))
row += 1
num_cols2 = ['Age','Tenure_in_Months','Monthly_Charge','Total_Charges',
             'Total_Revenue','Total_Long_Distance_Charges','Number_of_Referrals']
for i, col in enumerate(num_cols2):
    fill = GRAY if i%2 else WHITE
    stats = df[col].describe()
    write(ws2, row, 1, col, bold=True, fill=fill)
    write(ws2, row, 2, int(stats['count']), '#,##0', fill=fill, align='right')
    write(ws2, row, 3, round(stats['mean'],2), '#,##0.00', fill=fill, align='right')
    write(ws2, row, 4, round(stats['std'],2),  '#,##0.00', fill=fill, align='right')
    write(ws2, row, 5, round(stats['min'],2),  '#,##0.00', fill=fill, align='right')
    write(ws2, row, 6, round(stats['50%'],2),  '#,##0.00', fill=fill, align='right')
    write(ws2, row, 7, round(stats['max'],2),  '#,##0.00', fill=fill, align='right')
    row += 1

# Missing values
row += 1
hdr_row(ws2, row, 1, 'Missing Values Summary', span=3)
row += 1
for h, col in zip(['Column','Missing Count','Missing %'], [1,2,3]):
    hdr_row(ws2, row, col, h, fill=PatternFill('solid', fgColor='475569'))
row += 1
miss = df.isnull().sum()
miss = miss[miss>0].sort_values(ascending=False)
for i, (col_name, cnt) in enumerate(miss.items()):
    fill = GRAY if i%2 else WHITE
    write(ws2, row, 1, col_name, bold=True, fill=fill)
    write(ws2, row, 2, cnt, '#,##0', fill=fill, align='right')
    write(ws2, row, 3, cnt/len(df), '0.0%', fill=fill, align='right')
    row += 1

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3: CHURN DEEP DIVE
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet('Churn Deep Dive')
ws3.column_dimensions['A'].width = 36
for c in 'BCD': ws3.column_dimensions[c].width = 18
ws3.row_dimensions[1].height = 40

ws3.merge_cells('A1:D1')
t = ws3['A1']; t.value = 'Churn Analysis – Deep Dive'
t.font = Font(name='Arial', bold=True, size=15, color='FFFFFF')
t.fill = PatternFill('solid', fgColor='1E3A5F')
t.alignment = Alignment(horizontal='center', vertical='center')

row = 3
hdr_row(ws3, row, 1, 'Top 15 Churn Reasons', span=3)
row += 1
for h, c in zip(['Reason','Count','% of Churned'], [1,2,3]):
    hdr_row(ws3, row, c, h, fill=PatternFill('solid', fgColor='475569'))
row += 1
for i, (reason, cnt) in enumerate(df_churned['Churn_Reason'].value_counts().head(15).items()):
    fill = GRAY if i%2 else WHITE
    write(ws3, row, 1, reason, fill=fill)
    write(ws3, row, 2, cnt, '#,##0', fill=fill, align='right')
    write(ws3, row, 3, cnt/len(df_churned), '0.0%', fill=fill, align='right')
    row += 1

row += 1
hdr_row(ws3, row, 1, 'Churn Rate by Internet Type', span=3)
row += 1
for h, c in zip(['Internet Type','Churned','Churn Rate'], [1,2,3]):
    hdr_row(ws3, row, c, h, fill=PatternFill('solid', fgColor='475569'))
row += 1
for i, itype in enumerate(df['Internet_Type'].dropna().unique()):
    sub = df[df['Internet_Type']==itype]
    ch  = (sub['Customer_Status']=='Churned').sum()
    fill = GRAY if i%2 else WHITE
    write(ws3, row, 1, itype, bold=True, fill=fill)
    write(ws3, row, 2, ch, '#,##0', fill=fill, align='right')
    write(ws3, row, 3, ch/len(sub), '0.0%', fill=fill, align='right')
    row += 1

row += 1
hdr_row(ws3, row, 1, 'Churn Rate by Payment Method', span=3)
row += 1
for h, c in zip(['Payment Method','Churned','Churn Rate'], [1,2,3]):
    hdr_row(ws3, row, c, h, fill=PatternFill('solid', fgColor='475569'))
row += 1
for i, pm in enumerate(df['Payment_Method'].unique()):
    sub = df[df['Payment_Method']==pm]
    ch  = (sub['Customer_Status']=='Churned').sum()
    fill = GRAY if i%2 else WHITE
    write(ws3, row, 1, pm, bold=True, fill=fill)
    write(ws3, row, 2, ch, '#,##0', fill=fill, align='right')
    write(ws3, row, 3, ch/len(sub), '0.0%', fill=fill, align='right')
    row += 1

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4: CHARTS (embed PNGs)
# ─────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet('Charts')
ws4.row_dimensions[1].height = 40
ws4.merge_cells('A1:P1')
t = ws4['A1']; t.value = 'EDA Visualizations'
t.font = Font(name='Arial', bold=True, size=15, color='FFFFFF')
t.fill = PatternFill('solid', fgColor='1E3A5F')
t.alignment = Alignment(horizontal='center', vertical='center')

chart_files = [
    (f'{IMG}/01_status_donut.png',   'A3'),
    (f'{IMG}/02_churn_category.png', 'I3'),
    (f'{IMG}/03_age_distribution.png','A23'),
    (f'{IMG}/04_charge_vs_tenure.png','I23'),
    (f'{IMG}/05_contract_churn.png', 'A43'),
    (f'{IMG}/06_churn_reasons.png',  'I43'),
    (f'{IMG}/07_internet_churn.png', 'A63'),
    (f'{IMG}/08_revenue_boxplot.png','I63'),
    (f'{IMG}/09_correlation.png',    'A83'),
    (f'{IMG}/10_state_churn.png',    'I83'),
]
for path, anchor in chart_files:
    img = XLImage(path)
    img.width = 480; img.height = 300
    ws4.add_image(img, anchor)

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 5: STATE ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet('State Analysis')
for c, w in zip('ABCDE', [28,14,14,14,14]): ws5.column_dimensions[c].width = w
ws5.row_dimensions[1].height = 40
ws5.merge_cells('A1:E1')
t = ws5['A1']; t.value = 'State-wise Customer & Churn Analysis'
t.font = Font(name='Arial', bold=True, size=15, color='FFFFFF')
t.fill = PatternFill('solid', fgColor='1E3A5F')
t.alignment = Alignment(horizontal='center', vertical='center')

row = 3
for h, c in zip(['State','Total Customers','Churned','Churn Rate','Avg Monthly Charge'], range(1,6)):
    hdr_row(ws5, row, c, h, fill=PatternFill('solid', fgColor='475569'))
row += 1
state_stats = df.groupby('State').agg(
    Total=('Customer_ID','count'),
    Churned=('Customer_Status', lambda x: (x=='Churned').sum()),
    Avg_Charge=('Monthly_Charge','mean')
).reset_index().sort_values('Total', ascending=False)
for i, r in state_stats.iterrows():
    fill = GRAY if i%2==0 else WHITE
    write(ws5, row, 1, r['State'], bold=True, fill=fill)
    write(ws5, row, 2, r['Total'], '#,##0', fill=fill, align='right')
    write(ws5, row, 3, r['Churned'], '#,##0', fill=fill, align='right')
    write(ws5, row, 4, r['Churned']/r['Total'], '0.0%', fill=fill, align='right')
    write(ws5, row, 5, round(r['Avg_Charge'],2), '$#,##0.00', fill=fill, align='right')
    row += 1

XLSX_PATH = f'{OUT}/Customer_EDA_Report.xlsx'
wb.save(XLSX_PATH)
print(f"Excel saved → {XLSX_PATH} ✓")


# ════════════════════════════════════════════════════════════════════════════
# PDF REPORT
# ════════════════════════════════════════════════════════════════════════════
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, PageBreak, Image as RLImage,
                                 HRFlowable)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

PDF_PATH = f'{OUT}/Customer_EDA_Report.pdf'
doc = SimpleDocTemplate(PDF_PATH, pagesize=A4,
                         leftMargin=2*cm, rightMargin=2*cm,
                         topMargin=2.5*cm, bottomMargin=2*cm)

W, H = A4
TW = W - 4*cm   # text width

styles = getSampleStyleSheet()
NAVY  = colors.HexColor('#1E3A5F')
BLUE  = colors.HexColor('#2563EB')
LGRAY = colors.HexColor('#F1F5F9')
RED   = colors.HexColor('#EF4444')
GREEN = colors.HexColor('#10B981')

title_style = ParagraphStyle('Title', parent=styles['Title'],
    fontSize=22, textColor=NAVY, spaceAfter=6, alignment=TA_CENTER)
h1 = ParagraphStyle('H1', parent=styles['Heading1'],
    fontSize=14, textColor=NAVY, spaceBefore=14, spaceAfter=6)
h2 = ParagraphStyle('H2', parent=styles['Heading2'],
    fontSize=11, textColor=BLUE, spaceBefore=8, spaceAfter=4)
body = ParagraphStyle('Body', parent=styles['Normal'],
    fontSize=9.5, leading=14, spaceAfter=4)
caption = ParagraphStyle('Caption', parent=styles['Normal'],
    fontSize=8.5, textColor=colors.grey, alignment=TA_CENTER, spaceAfter=8)

def tbl_style(header_color=NAVY):
    return TableStyle([
        ('BACKGROUND', (0,0), (-1,0), header_color),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 9),
        ('ALIGN',      (0,0), (-1,0), 'CENTER'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, LGRAY]),
        ('FONTNAME',   (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',   (0,1), (-1,-1), 8.5),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#CBD5E1')),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0),(-1,-1), 4),
    ])

story = []

# ── Cover Page ────────────────────────────────────────────────────────────────
story.append(Spacer(1, 3*cm))
story.append(Paragraph('Customer Data', ParagraphStyle('sub', parent=styles['Normal'],
    fontSize=13, textColor=BLUE, alignment=TA_CENTER)))
story.append(Spacer(1, 0.4*cm))
story.append(Paragraph('Exploratory Data Analysis Report',
    ParagraphStyle('cover', parent=styles['Title'], fontSize=26,
                   textColor=NAVY, alignment=TA_CENTER)))
story.append(HRFlowable(width=TW, thickness=2, color=BLUE, spaceAfter=20))
story.append(Spacer(1, 0.5*cm))

kpi_data = [
    ['Metric', 'Value'],
    ['Total Customers',    f"{len(df):,}"],
    ['Churned Customers',  f"{len(df_churned):,}"],
    ['Overall Churn Rate', f"{churn_rate}%"],
    ['Avg Monthly Charge', f"${df['Monthly_Charge'].mean():.2f}"],
    ['Avg Tenure',         f"{df['Tenure_in_Months'].mean():.1f} months"],
    ['Total Revenue',      f"${df['Total_Revenue'].sum():,.0f}"],
]
t = Table(kpi_data, colWidths=[TW*0.55, TW*0.45])
t.setStyle(tbl_style())
story.append(t)
story.append(PageBreak())

# ── Section 1: Dataset Overview ───────────────────────────────────────────────
story.append(Paragraph('1. Dataset Overview', h1))
story.append(HRFlowable(width=TW, thickness=1, color=BLUE, spaceAfter=8))
story.append(Paragraph(
    f'The dataset contains <b>{len(df):,} customer records</b> across <b>{len(df.columns)} attributes</b>. '
    f'It covers demographics, service subscriptions, billing information, and customer status. '
    f'Of all customers, <b>{len(df_churned):,} ({churn_rate}%)</b> have churned, '
    f'<b>{(df["Customer_Status"]=="Stayed").sum():,}</b> stayed, and '
    f'<b>{(df["Customer_Status"]=="Joined").sum():,}</b> are newly joined.',
    body))

story.append(Paragraph('1.1 Customer Status Distribution', h2))
img = RLImage(f'{IMG}/01_status_donut.png', width=TW*0.55, height=TW*0.38)
img.hAlign = 'CENTER'; story.append(img)
story.append(Paragraph('Figure 1: Customer Status Distribution (Donut Chart)', caption))

story.append(Paragraph('1.2 Descriptive Statistics', h2))
desc_data = [['Column','Mean','Std Dev','Min','Median','Max']]
for col in ['Age','Tenure_in_Months','Monthly_Charge','Total_Revenue']:
    s = df[col].describe()
    desc_data.append([col, f'{s["mean"]:.1f}', f'{s["std"]:.1f}',
                      f'{s["min"]:.1f}', f'{s["50%"]:.1f}', f'{s["max"]:.1f}'])
t = Table(desc_data, colWidths=[TW*0.28]+[TW*0.144]*5)
t.setStyle(tbl_style())
story.append(t)
story.append(PageBreak())

# ── Section 2: Churn Analysis ─────────────────────────────────────────────────
story.append(Paragraph('2. Churn Analysis', h1))
story.append(HRFlowable(width=TW, thickness=1, color=BLUE, spaceAfter=8))
story.append(Paragraph(
    f'Churn analysis is the core focus of this report. The overall churn rate stands at '
    f'<b>{churn_rate}%</b>. The dominant churn driver is <b>Competitor-related reasons</b> '
    f'({df_churned["Churn_Category"].value_counts().iloc[0]:,} customers), followed by '
    f'<b>Attitude</b> and <b>Dissatisfaction</b>.', body))

story.append(Paragraph('2.1 Churn by Category', h2))
img = RLImage(f'{IMG}/02_churn_category.png', width=TW*0.8, height=TW*0.45)
img.hAlign = 'CENTER'; story.append(img)
story.append(Paragraph('Figure 2: Churn Reasons by Category', caption))

story.append(Paragraph('2.2 Top 10 Specific Churn Reasons', h2))
top10 = df_churned['Churn_Reason'].value_counts().head(10)
reasons_data = [['Rank','Churn Reason','Count','% of Churned']]
for i, (reason, cnt) in enumerate(top10.items(), 1):
    reasons_data.append([str(i), reason, f'{cnt:,}', f'{cnt/len(df_churned)*100:.1f}%'])
t = Table(reasons_data, colWidths=[TW*0.07, TW*0.55, TW*0.18, TW*0.20])
t.setStyle(tbl_style())
story.append(t)
story.append(Spacer(1, 0.4*cm))

story.append(Paragraph('2.3 Top 10 Churn Reasons – Chart', h2))
img = RLImage(f'{IMG}/06_churn_reasons.png', width=TW*0.85, height=TW*0.52)
img.hAlign = 'CENTER'; story.append(img)
story.append(Paragraph('Figure 3: Top 10 Churn Reasons', caption))
story.append(PageBreak())

# ── Section 3: Demographics ───────────────────────────────────────────────────
story.append(Paragraph('3. Demographic Analysis', h1))
story.append(HRFlowable(width=TW, thickness=1, color=BLUE, spaceAfter=8))
story.append(Paragraph(
    f'The customer base spans ages <b>18 to 85</b> with a mean of '
    f'<b>{df["Age"].mean():.1f} years</b>. Churned customers tend to be slightly '
    f'older on average ({df_churned["Age"].mean():.1f} years) compared to those who '
    f'stayed ({df[df["Customer_Status"]=="Stayed"]["Age"].mean():.1f} years).', body))

img = RLImage(f'{IMG}/03_age_distribution.png', width=TW*0.9, height=TW*0.5)
img.hAlign = 'CENTER'; story.append(img)
story.append(Paragraph('Figure 4: Age Distribution by Customer Status', caption))

story.append(Paragraph('3.1 Gender Split', h2))
gender_data = [['Gender','Count','Churn Count','Churn Rate']]
for g in df['Gender'].unique():
    sub = df[df['Gender']==g]
    ch  = (sub['Customer_Status']=='Churned').sum()
    gender_data.append([g, f'{len(sub):,}', f'{ch:,}', f'{ch/len(sub)*100:.1f}%'])
t = Table(gender_data, colWidths=[TW*0.25]*4)
t.setStyle(tbl_style())
story.append(t)
story.append(PageBreak())

# ── Section 4: Service & Contract ─────────────────────────────────────────────
story.append(Paragraph('4. Service & Contract Analysis', h1))
story.append(HRFlowable(width=TW, thickness=1, color=BLUE, spaceAfter=8))

story.append(Paragraph('4.1 Contract Type vs Churn', h2))
img = RLImage(f'{IMG}/05_contract_churn.png', width=TW*0.75, height=TW*0.48)
img.hAlign = 'CENTER'; story.append(img)
story.append(Paragraph('Figure 5: Churn Rate by Contract Type', caption))
story.append(Paragraph(
    'Month-to-Month contracts have the highest churn rate, indicating that longer-term '
    'contracts are strongly associated with customer retention.', body))

story.append(Paragraph('4.2 Internet Type vs Churn', h2))
img = RLImage(f'{IMG}/07_internet_churn.png', width=TW*0.75, height=TW*0.48)
img.hAlign = 'CENTER'; story.append(img)
story.append(Paragraph('Figure 6: Churn Rate by Internet Type', caption))
story.append(PageBreak())

# ── Section 5: Revenue ────────────────────────────────────────────────────────
story.append(Paragraph('5. Revenue & Billing Analysis', h1))
story.append(HRFlowable(width=TW, thickness=1, color=BLUE, spaceAfter=8))

story.append(Paragraph('5.1 Monthly Charge vs Tenure', h2))
img = RLImage(f'{IMG}/04_charge_vs_tenure.png', width=TW*0.9, height=TW*0.55)
img.hAlign = 'CENTER'; story.append(img)
story.append(Paragraph('Figure 7: Monthly Charge vs Tenure by Status', caption))

story.append(Paragraph('5.2 Revenue Distribution by Status', h2))
img = RLImage(f'{IMG}/08_revenue_boxplot.png', width=TW*0.9, height=TW*0.5)
img.hAlign = 'CENTER'; story.append(img)
story.append(Paragraph('Figure 8: Monthly Charge & Total Revenue Boxplots by Status', caption))
story.append(PageBreak())

# ── Section 6: Correlation ────────────────────────────────────────────────────
story.append(Paragraph('6. Correlation Analysis', h1))
story.append(HRFlowable(width=TW, thickness=1, color=BLUE, spaceAfter=8))
story.append(Paragraph(
    'The heatmap below shows Pearson correlations between numeric features. '
    'Total_Charges and Total_Revenue are highly correlated (expected). '
    'Tenure is moderately positively correlated with Total_Charges.', body))
img = RLImage(f'{IMG}/09_correlation.png', width=TW*0.9, height=TW*0.65)
img.hAlign = 'CENTER'; story.append(img)
story.append(Paragraph('Figure 9: Correlation Heatmap', caption))
story.append(PageBreak())

# ── Section 7: Geographic ─────────────────────────────────────────────────────
story.append(Paragraph('7. Geographic Analysis', h1))
story.append(HRFlowable(width=TW, thickness=1, color=BLUE, spaceAfter=8))
img = RLImage(f'{IMG}/10_state_churn.png', width=TW*0.9, height=TW*0.55)
img.hAlign = 'CENTER'; story.append(img)
story.append(Paragraph('Figure 10: Top 10 States by Churn Rate', caption))

top5_states = df.groupby('State').apply(
    lambda x: (x['Customer_Status']=='Churned').sum()/len(x)*100
).sort_values(ascending=False).head(5)
state_tbl = [['State','Churn Rate']]
for s, r in top5_states.items():
    state_tbl.append([s, f'{r:.1f}%'])
t = Table(state_tbl, colWidths=[TW*0.5, TW*0.5])
t.setStyle(tbl_style())
story.append(t)
story.append(PageBreak())

# ── Section 8: Key Findings & Recommendations ─────────────────────────────────
story.append(Paragraph('8. Key Findings & Recommendations', h1))
story.append(HRFlowable(width=TW, thickness=1, color=BLUE, spaceAfter=8))

findings = [
    ('High Churn Rate',
     f'Overall churn is {churn_rate}%, predominantly driven by competitor offers. '
     'Competitive pricing and service improvements are critical.'),
    ('Month-to-Month Risk',
     'Customers on Month-to-Month contracts churn at a significantly higher rate. '
     'Incentivizing migration to 1- or 2-year contracts can reduce churn.'),
    ('Fiber Optic Churn',
     'Fiber Optic internet users show higher churn, suggesting service quality or '
     'pricing issues that need targeted attention.'),
    ('Revenue Impact',
     'Churned customers had higher average monthly charges, meaning revenue loss '
     'per churned customer is above average.'),
    ('Geographic Hotspots',
     'Certain states show disproportionately high churn rates — targeted retention '
     'campaigns in those regions could deliver high ROI.'),
    ('Tenure & Loyalty',
     'Customers with longer tenure churn less. Early-tenure engagement programs '
     '(first 6 months) could significantly reduce early churn.'),
]
for title_str, detail in findings:
    story.append(Paragraph(f'<b>{title_str}</b>', h2))
    story.append(Paragraph(detail, body))
    story.append(Spacer(1, 0.2*cm))

story.append(Spacer(1, 1*cm))
story.append(HRFlowable(width=TW, thickness=1, color=colors.HexColor('#CBD5E1'), spaceAfter=6))
story.append(Paragraph('Report generated via automated EDA pipeline. Data: Customer_Data.csv',
    ParagraphStyle('footer', parent=styles['Normal'], fontSize=8,
                   textColor=colors.grey, alignment=TA_CENTER)))

doc.build(story)
print(f"PDF saved → {PDF_PATH} ✓")
print("\nAll outputs complete.")
